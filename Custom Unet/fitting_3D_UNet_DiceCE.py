import torch
import numpy as np
import h5py
from torch.utils.data import DataLoader
from monai.losses import DiceCELoss
import openpyxl

device = "cuda:2"
lossDevice = "cuda:3"
model_name = "trained_3DUnetDiceCE.pth"

class CustomDataset(torch.utils.data.Dataset):
    def __init__(self, images, masks):
        self.images = images
        self.masks = masks

        with h5py.File(self.images, mode='r') as f:
            self.file_names = list(f.keys())

    def __len__(self):
        return len(self.file_names)
    
    def __getitem__(self, idx):
        file_name = self.file_names[idx]
        with h5py.File(self.images, mode='r') as i:
            image = i[file_name][:]
        with h5py.File(self.masks, mode='r') as m:
            mask = m[file_name][:]
        image = np.expand_dims(image, axis=0)
        mask = np.expand_dims(mask, axis=0)
        image = torch.from_numpy(image)
        mask = torch.from_numpy(mask.astype(np.uint8))

        return image, mask

def train_loop(dataloader, model, loss_function, optimizer):
    size = len(dataloader.dataset)
    train_loss = 0

    model.train()
    for batch, (x, y) in enumerate(dataloader):
        x = x.to(device)
        y = y.to(lossDevice)

        pred = model(x).to(lossDevice)
        loss = loss_function(pred, y)
        train_loss += loss.item()

        loss.backward()
        optimizer.step()
        optimizer.zero_grad()
        if batch % 10 == 0:
            current = (batch + 1) * len(x)
            avg_loss = train_loss / current
            print(f"Average train loss: {avg_loss:>7f} Instance train loss: {loss.item():>7f} [{current:>5d}/{size:>5d}]")

    return train_loss / size

def test_loop(dataloader, model, loss_function):
    model.eval()
    size = len(dataloader.dataset)
    test_loss = 0

    with torch.no_grad():
        for x, y in dataloader:
            x = x.to(device)
            y = y.to(lossDevice)
            pred = model(x).to(lossDevice)
            test_loss += loss_function(pred, y).item()
    
    test_loss /= size
    print(f"Test Error: Avg loss: {test_loss:>8f} \n")

    return test_loss

def save_model(loss, model, current_min):
    if loss < current_min:
        torch.save(model, model_name)
        print(f"New min test_loss of {loss} \n")
        return loss
    
    return current_min


def add_losses_to_excel(filename, train_losses, test_losses):
    # Load the existing Excel file or create a new one if it doesn't exist
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Select the first sheet (assuming it's the default one)
    sheet = workbook.active

    # Find the last row in the sheet
    last_row = sheet.max_row + 1

    # Add train and test losses to the sheet
    for train_loss, test_loss in zip(train_losses, test_losses):
        sheet.cell(row=last_row, column=1, value=train_loss)
        sheet.cell(row=last_row, column=2, value=test_loss)
        last_row += 1

    # Save the changes to the Excel file
    workbook.save(filename)


if __name__ == "__main__":
    train_data = CustomDataset("WORD_Tr_img.h5", "WORD_Tr_mask.h5")
    test_data = CustomDataset("WORD_Val_img.h5", "WORD_Val_mask.h5")

    train_dataloader = DataLoader(train_data, batch_size=1, shuffle=True)
    test_dataloader = DataLoader(test_data, batch_size=1, shuffle=True)

    # from monai.networks.nets import UNet
    # model = UNet(
    #     3,
    #     in_channels=1,
    #     out_channels=17,
    #     channels=(16, 32, 64, 128, 256), 
    #     strides=(2, 2, 2, 2)
    # ).model

    model = torch.load(model_name)
    model.to(device)
    learning_rate = 1e-3
    weight_decay = 1e-5
    loss_function = DiceCELoss(include_background=True, to_onehot_y=True, sigmoid=True, squared_pred=True)
    optimizer = torch.optim.Adam(model.parameters(), learning_rate, weight_decay=weight_decay, amsgrad=True)

    epochs = 150
    min = 0.30801132
    train_loss = list()
    test_loss = list()
    for t in range(epochs):
        print(f"Epoch {t+1}\n-------------------------------")
        train_loss.append(train_loop(train_dataloader, model, loss_function, optimizer))
        t_loss = test_loop(test_dataloader, model, loss_function)
        test_loss.append(t_loss)
        min = save_model(t_loss, model, min)
    print(f"Done! Min test_loss was {min}")

    add_losses_to_excel("3D UNet Training Log DiceCE Loss.xlsx", train_loss, test_loss)
